from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time

wb = load_workbook('dzivokli.xlsx')
ws = wb['Filtrs']
ws2 = wb['Sludinajumi']

max_row = ws.max_row

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

# Nolasa filtru vērtības no Excel faila
for row in range(2, max_row+1):
    lokacija = (ws['a' + str(row)].value)
    max_cena = (ws['b' + str(row)].value)
    dar_veids = (ws['c' + str(row)].value)

url = "https://www.ss.com/lv/real-estate/flats/" + lokacija
driver.get(url)

# Izvēlas darījuma veidu
select_element = driver.find_element(By.NAME, "sid")
select = Select(select_element)
select.select_by_visible_text(dar_veids)

# Uzliek sludinājumu maksimālo cenu
element = driver.find_element(By.ID, "f_o_8_max")
element.send_keys(max_cena)

# Nospiež pogu meklēt
meklet = driver.find_element(
    By.CSS_SELECTOR, "input[type='submit'][value='Meklēt']")
meklet.click()

time.sleep(5)

# Atlasa pēdējos 15 sludinājumus
post_rows = driver.find_elements(
    By.XPATH, "//tr[contains(@id, 'tr_')]")[:15]
sludinajumi = []
for row in post_rows:
    iela = row.find_element(
        By.XPATH, ".//td[contains(@class, 'msga2-o')][1]/a").text

    m2 = row.find_element(
        By.XPATH, ".//td[contains(@class, 'msga2-o')][3]/a").text
    cena = row.find_element(
        By.XPATH, ".//td[contains(@class, 'msga2-o')][last()]/a").text

    link = row.find_element(
        By.XPATH, ".//td[contains(@class, 'msg2')]/div/a").get_attribute('href')

    sludinajumi.append(
        {"iela": iela, "m2": m2, "cena": cena, "link": link})

row = 2

# Atlasītos sludinājumus saglabā Excel failā
for s in sludinajumi:
    ws2.cell(row=row, column=1, value=s['iela'])
    ws2.cell(row=row, column=2, value=int(s['m2']))
    ws2.cell(row=row, column=3, value=s['cena'])
    ws2.cell(row=row, column=4, value=s['link'])
    row += 1

wb.save('dzivokli.xlsx')
driver.close()
